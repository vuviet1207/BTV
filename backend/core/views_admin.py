# after
import csv
import unicodedata 
import re          
from io import TextIOWrapper
from django.contrib import messages
from django.shortcuts import render, redirect
from django.db import transaction
from django.core.files.uploadedfile import InMemoryUploadedFile, TemporaryUploadedFile
from openpyxl import load_workbook
from django.db.models import Prefetch
from .models import BaiThi, BaiThiTemplateSection, BaiThiTemplateItem, VongThi
from .models import ThiSinh, GiamKhao, CuocThi, ThiSinhCuocThi
from core.decorators import judge_required

import os
from django.conf import settings
from django.core.files.storage import default_storage

REQUIRED_COLUMNS = {
    "thisinh": ["maNV", "hoTen", "chiNhanh", "vung", "donVi", "email", "nhom", "image_url"],
    "giamkhao": ["maNV", "hoTen", "email"],
}

# ===== Alias & chu·∫©n h√≥a header =====
def _normalize(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ascii", "ignore").decode("ascii")  # b·ªè d·∫•u
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "", s)  # b·ªè k√Ω t·ª± l·∫° & kho·∫£ng tr·∫Øng
    return s

HEADER_ALIASES = {
    # maNV
    "manv": "maNV", "manhanvien": "maNV", "manvnv": "maNV", "ma": "maNV", "ma_nv": "maNV",
    # hoTen
    "hoten": "hoTen", "ten": "hoTen", "hovaten": "hoTen", "ho_ten": "hoTen",
    # chiNhanh
    "chinhanh": "chiNhanh", "chi_nhanh": "chiNhanh", "cn": "chiNhanh",
    # vung
    "vung": "vung", "mien": "vung",
    # donVi
    "donvi": "donVi", "don_vi": "donVi", "dv": "donVi", "don": "donVi", "donvichuyendoi": "donVi",
    # email
    "email": "email", "mail": "email", "e-mail": "email",
    # nhom
    "nhom": "nhom", "group": "nhom", "nhomthi": "nhom",
    #imageUrl
    "imageurl": "image_url", "image_url": "image_url", "hinhanh": "image_url",
    "hinh_anh": "image_url", "hinhAnh": "image_url", "anh": "image_url", "img": "image_url"

}

def _map_header_list(header, expected_cols):
    """
    Tr·∫£ v·ªÅ: (canon_order, source_idx)
    canon_order: danh s√°ch t√™n c·ªôt ƒë√£ ƒë∆∞·ª£c map v·ªÅ canonical (theo th·ª© t·ª± header g·ªëc)
    source_idx: dict {canonical_name: index_goc}
    """
    canon_order = []
    for h in header:
        key = _normalize(h or "")
        canon = HEADER_ALIASES.get(key)
        canon_order.append(canon or (h or "").strip())

    src_idx = {}
    for i, canon in enumerate(canon_order):
        if canon not in src_idx:
            src_idx[canon] = i

    missing = [c for c in expected_cols if c not in src_idx]
    return canon_order, src_idx, missing

# after
def _read_xlsx(file, expected_cols):
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    _, src_idx, missing = _map_header_list(header, expected_cols)
    if missing:
        raise ValueError(f"Thi·∫øu c·ªôt: {', '.join(missing)}")

    data = []
    for r in rows[1:]:
        if r is None:
            continue
        row = {}
        for c in expected_cols:
            idx = src_idx[c]
            val = r[idx] if idx < len(r) else None
            row[c] = "" if val is None else str(val).strip()
        data.append(row)
    return data

# after
def _read_csv(file, expected_cols):
    # file l√† UploadedFile -> c·∫ßn decode UTF-8
    text_stream = TextIOWrapper(file, encoding="utf-8")
    reader = csv.DictReader(text_stream)
    header = reader.fieldnames or []

    # map header g·ªëc -> canonical
    canon_order, src_idx, missing = _map_header_list(header, expected_cols)
    if missing:
        raise ValueError(f"Thi·∫øu c·ªôt: {', '.join(missing)}")

    # X√¢y map canonical -> t√™n c·ªôt g·ªëc ƒë·ªÉ l·∫•y d·ªØ li·ªáu
    # (∆∞u ti√™n ƒë√∫ng ch·ªØ canonical n·∫øu ƒë√£ c√≥, n·∫øu kh√¥ng l·∫•y t·ª´ alias)
    canon_to_source = {}
    for i, h in enumerate(header):
        key = _normalize(h or "")
        canon = HEADER_ALIASES.get(key) or (h or "").strip()
        if canon not in canon_to_source:
            canon_to_source[canon] = h

    data = []
    for row in reader:
        out = {}
        for c in expected_cols:
            src = canon_to_source.get(c, c)
            out[c] = (row.get(src, "") or "").strip()
        data.append(out)
    return data
def _find_duplicate_ma_email(rows, key_ma="maNV", key_email="email"):
    """
    rows: list[dict] ƒë·ªçc t·ª´ file import
    Tr·∫£ v·ªÅ:
      - dup_ma: set c√°c m√£ NV tr√πng trong file
      - dup_email: set c√°c email tr√πng trong file
    """
    seen_ma = set()
    seen_email = set()
    dup_ma = set()
    dup_email = set()

    for r in rows:
        ma = (r.get(key_ma) or "").strip()
        if ma:
            if ma in seen_ma:
                dup_ma.add(ma)
            else:
                seen_ma.add(ma)

        email = (r.get(key_email) or "").strip().lower()
        if email:
            if email in seen_email:
                dup_email.add(email)
            else:
                seen_email.add(email)

    return dup_ma, dup_email

@judge_required
def import_view(request):
    preselected_ma = None
    q = request.GET.get("ct") or request.POST.get("maCT")  # c·∫£ GET & POST
    if q:
        # n·∫øu q l√† m√£ (CT001...)
        if CuocThi.objects.filter(ma=q).exists():
            preselected_ma = q
        else:
            # n·∫øu q l√† id s·ªë -> chuy·ªÉn sang m√£
            try:
                obj = CuocThi.objects.only("ma").get(pk=int(q))
                preselected_ma = obj.ma
            except Exception:
                pass
    if request.method == "POST":
        target = request.POST.get("target")  # thisinh | giamkhao
        selected_ma_ct = request.POST.get("maCT")  # NEW
        f = request.FILES.get("file")

        cuocthi_obj = None
        if selected_ma_ct:
            cuocthi_obj = CuocThi.objects.filter(ma=selected_ma_ct).first()

        if target not in REQUIRED_COLUMNS:
            messages.error(request, "Vui l√≤ng ch·ªçn lo·∫°i d·ªØ li·ªáu h·ª£p l·ªá.")
            return redirect(request.path)
        if not f:
            messages.error(request, "Vui l√≤ng ch·ªçn t·ªáp CSV/XLSX.")
            return redirect(request.path)

        expected = REQUIRED_COLUMNS[target]
        try:
            if isinstance(f, (InMemoryUploadedFile, TemporaryUploadedFile)) and f.name.lower().endswith(".xlsx"):
                rows = _read_xlsx(f, expected)
            else:
                rows = _read_csv(f, expected)
        except Exception as e:
            messages.error(request, f"L·ªói ƒë·ªçc t·ªáp: {e}")
            return redirect(request.path)

        # üî¥ NEW: ki·ªÉm tra tr√πng m√£ / email trong file
        dup_ma, dup_email = _find_duplicate_ma_email(rows)
        if dup_ma or dup_email:
            # t√™n lo·∫°i d·ªØ li·ªáu ƒë·ªÉ hi·ªán cho d·ªÖ hi·ªÉu
            loai = "th√≠ sinh" if target == "thisinh" else "gi√°m kh·∫£o"
            parts = []
            if dup_ma:
                parts.append("M√£ nh√¢n vi√™n tr√πng: " + ", ".join(sorted(dup_ma)))
            if dup_email:
                parts.append("Email tr√πng: " + ", ".join(sorted(dup_email)))

            if cuocthi_obj:
                prefix = f"Kh√¥ng th·ªÉ import {loai} v√†o cu·ªôc thi {cuocthi_obj.ma} v√¨ t·ªáp c√≥ nhi·ªÅu d√≤ng tr√πng nhau. "
            else:
                prefix = f"Kh√¥ng th·ªÉ import {loai} v√¨ t·ªáp c√≥ nhi·ªÅu d√≤ng tr√πng nhau. "

            messages.error(request, prefix + " | ".join(parts))
            return redirect(request.path)
        # üî¥ H·∫øt ph·∫ßn check tr√πng

        created = updated = skipped = 0
        with transaction.atomic():
            if target == "thisinh":
                for r in rows:
                    ma = r.get("maNV", "").strip()
                    if not ma:
                        skipped += 1
                        continue

                    # Required
                    hoTen = (r.get("hoTen", "") or "").strip()

                    # Optional fields: convert empty string -> None so DB stores NULL (avoids unique '' collisions on email)
                    chiNhanh = (r.get("chiNhanh", "") or "").strip() or None
                    vung = (r.get("vung", "") or "").strip() or None
                    donVi = (r.get("donVi", "") or "").strip() or None
                    email = (r.get("email", "") or "").strip() or None
                    nhom = (r.get("nhom", "") or "").strip() or None
                    image_url = (r.get("image_url", "") or "").strip() or None

                    obj, is_created = ThiSinh.objects.update_or_create(
                        pk=ma,
                        defaults=dict(
                            hoTen=hoTen,
                            chiNhanh=chiNhanh,
                            vung=vung,
                            donVi=donVi,
                            email=email,
                            nhom=nhom,
                            image_url=image_url,
                        ),
                    )

                    # ‚úÖ Ghi v√†o b·∫£ng tham gia n·∫øu c√≥ ch·ªçn cu·ªôc thi
                    if cuocthi_obj:
                        try:
                            ThiSinhCuocThi.objects.get_or_create(
                                thiSinh=obj, cuocThi=cuocthi_obj
                            )
                        except Exception as e:
                            messages.warning(request, f"L·ªói khi t·∫°o quan h·ªá cho {ma}: {e}")
                    created += int(is_created)
                    updated += int(not is_created)

            else:  # giamkhao
                for r in rows:
                    ma = r["maNV"]
                    if not ma:
                        skipped += 1
                        continue
                    # Ensure imported judges default to role 'JUDGE' when created/updated
                    obj, is_created = GiamKhao.objects.update_or_create(
                        pk=ma,
                        defaults=dict(
                            hoTen=r["hoTen"],
                            email=r["email"],
                            role="JUDGE",
                        )
                    )
                    created += int(is_created)
                    updated += int(not is_created)

        messages.success(request, f"Import xong: th√™m {created}, c·∫≠p nh·∫≠t {updated}, b·ªè qua {skipped}.")
        return redirect(request.path)


    return render(
        request,
        "importer/index.html",
        {
            "cuocthi_list": CuocThi.objects.all().values("ma", "tenCuocThi").order_by("ma"),
            "preselected_ma": preselected_ma,
        }
        
)

@judge_required
def upload_avatars_view(request):
    """
    Trang upload nhi·ªÅu ·∫£nh avatar.
    - T√™n file = maNV (kh√¥ng c·∫ßn ph√¢n bi·ªát hoa/th∆∞·ªùng), v√≠ d·ª•: NV001.jpg
    - L∆∞u v√†o MEDIA_ROOT/thisinh/maNV.jpg
    - T·ª± ƒë·ªông c·∫≠p nh·∫≠t ThiSinh.image_url t∆∞∆°ng ·ª©ng.
    """
    if request.method == "POST":
        files = request.FILES.getlist("images")
        if not files:
            messages.error(request, "Vui l√≤ng ch·ªçn √≠t nh·∫•t 1 ·∫£nh.")
            return redirect(request.path)

        updated = 0
        not_found = []
        skipped = 0

        for f in files:
            original_name = f.name or ""
            base_name = os.path.basename(original_name)
            stem, ext = os.path.splitext(base_name)

            ma = (stem or "").strip()
            ext = ext.lower()

            # Ch·ªâ nh·∫≠n jpg/jpeg/png
            if ext not in [".jpg", ".jpeg", ".png"]:
                skipped += 1
                continue
            if not ma:
                skipped += 1
                continue

            # T√¨m th√≠ sinh theo maNV (kh√¥ng ph√¢n bi·ªát hoa/th∆∞·ªùng)
            try:
                ts = ThiSinh.objects.get(maNV__iexact=ma)
            except ThiSinh.DoesNotExist:
                not_found.append(ma)
                continue

            # ƒê·∫∑t l·∫°i t√™n file chu·∫©n theo maNV trong DB (gi·ªØ nguy√™n hoa/th∆∞·ªùng trong DB)
            filename = f"{ts.maNV}{ext}"
            upload_path = os.path.join("thisinh", filename)

            # N·∫øu ƒë√£ c√≥ file c≈©, c√≥ th·ªÉ x√≥a ƒëi (tu·ª≥ nhu c·∫ßu)
            if default_storage.exists(upload_path):
                default_storage.delete(upload_path)

            # L∆∞u file m·ªõi
            saved_path = default_storage.save(upload_path, f)

            # L·∫•y URL ƒë·ªÉ l∆∞u v√†o image_url
            try:
                url = default_storage.url(saved_path)
            except Exception:
                # fallback: MEDIA_URL + path
                url = settings.MEDIA_URL + saved_path

            ts.image_url = url
            ts.save(update_fields=["image_url"])
            updated += 1

        if updated:
            messages.success(request, f"ƒê√£ c·∫≠p nh·∫≠t ·∫£nh cho {updated} th√≠ sinh.")
        if skipped:
            messages.warning(
                request,
                f"B·ªè qua {skipped} file kh√¥ng h·ª£p l·ªá (kh√¥ng ƒë√∫ng ƒë·ªãnh d·∫°ng jpg/png ho·∫∑c kh√¥ng c√≥ m√£)."
            )
        if not_found:
            messages.warning(
                request,
                "Kh√¥ng t√¨m th·∫•y th√≠ sinh cho c√°c m√£: " + ", ".join(sorted(set(not_found)))
            )

        return redirect(request.path)

    # GET: render trang upload
    return render(request, "importer/upload_avatars.html", {})



def organize_view(request):
    if request.method == "POST":
        action = request.POST.get("action")
        # ===== Import M·∫™U CH·∫§M cho 1 B√†i thi =====
        if action == "config_template_upload":
            btid = request.POST.get("baiThi_id")
            f = request.FILES.get("template_file")
            if not btid or not f:
                messages.error(request, "Thi·∫øu B√†i thi ho·∫∑c t·ªáp Excel.")
                return redirect(request.path)

            # L·∫•y b√†i thi
            try:
                bai_thi = BaiThi.objects.get(pk=btid)
            except BaiThi.DoesNotExist:
                messages.error(request, "Kh√¥ng t√¨m th·∫•y B√†i thi.")
                return redirect(request.path)

            # ƒê·ªçc Excel b·∫±ng openpyxl
            try:
                wb = load_workbook(f, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            except Exception as e:
                messages.error(request, f"L·ªói ƒë·ªçc Excel: {e}")
                return redirect(request.path)

            if not rows:
                messages.error(request, "T·ªáp r·ªóng.")
                return redirect(request.path)

            # Header & alias
            header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
            def idx(*names):
                for name in names:
                    if name in header:
                        return header.index(name)
                return None

            idx_section = idx("section", "m·ª•c l·ªõn", "muc lon", "phan", "ph·∫ßn")
            idx_item    = idx("item", "m·ª•c nh·ªè", "muc nho", "b√†i", "bai")
            idx_max     = idx("ƒëi·ªÉm", "diem", "max", "ƒëi·ªÉm t·ªëi ƒëa", "diem toi da")
            idx_note    = idx("note", "ghi ch√∫", "ghi chu")

            if idx_section is None or idx_max is None or (idx_item is None and idx_section is None):
                messages.error(request, "Thi·∫øu c·ªôt b·∫Øt bu·ªôc: Section/M·ª•c l·ªõn, Item/M·ª•c nh·ªè (n·∫øu c√≥), ƒêi·ªÉm/Max.")
                return redirect(request.path)

            # X√≥a template c≈©
            BaiThiTemplateSection.objects.filter(baiThi=bai_thi).delete()

            # Parse & L∆∞u m·ªõi theo th·ª© t·ª± xu·∫•t hi·ªán
            section_map = {}   # title -> (section_obj, next_item_stt)
            next_section_stt = 1
            created_items = 0

            for r in rows[1:]:
                if r is None:
                    continue
                sec_title = (r[idx_section] if idx_section is not None and idx_section < len(r) else None)
                item_text = (r[idx_item] if idx_item is not None and idx_item < len(r) else None)
                max_val   = (r[idx_max] if idx_max is not None and idx_max < len(r) else None)
                note_text = (r[idx_note] if idx_note is not None and idx_note < len(r) else None)

                # B·ªè d√≤ng tr·ªëng ho√†n to√†n
                if not (sec_title or item_text or max_val or note_text):
                    continue

                sec_title = str(sec_title).strip() if sec_title is not None else ""
                item_text = str(item_text).strip() if item_text is not None else ""
                note_text = str(note_text).strip() if note_text is not None else ""

                # Chu·∫©n ho√° ƒëi·ªÉm t·ªëi ƒëa
                try:
                    max_score = int(float(max_val)) if max_val is not None and str(max_val).strip() != "" else 0
                except Exception:
                    messages.error(request, f"ƒêi·ªÉm t·ªëi ƒëa kh√¥ng h·ª£p l·ªá ·ªü d√≤ng c√≥ m·ª•c: '{item_text or sec_title}'.")
                    return redirect(request.path)

                # T·∫°o/l·∫•y Section
                if sec_title not in section_map:
                    s = BaiThiTemplateSection.objects.create(
                        baiThi=bai_thi, stt=next_section_stt, title=sec_title or "M·ª•c"
                    )
                    section_map[sec_title] = [s, 1]  # (section_obj, next_item_stt)
                    next_section_stt += 1
                else:
                    s, _ = section_map[sec_title]

                # C√≥ Item (l√°) -> t·∫°o item d∆∞·ªõi section
                if item_text:
                    s, next_item_stt = section_map[sec_title]
                    BaiThiTemplateItem.objects.create(
                        section=s, stt=next_item_stt, content=item_text, max_score=max_score, note=note_text or None
                    )
                    section_map[sec_title][1] = next_item_stt + 1
                    created_items += 1
                else:
                    # Kh√¥ng c√≥ Item => hi·ªÉu l√† "nh√°nh kh√¥ng c√≥ l√°": l∆∞u 1 item ƒë·∫∑c bi·ªát
                    s, next_item_stt = section_map[sec_title]
                    BaiThiTemplateItem.objects.create(
                        section=s, stt=next_item_stt, content=s.title, max_score=max_score, note=note_text or None
                    )
                    section_map[sec_title][1] = next_item_stt + 1
                    created_items += 1

            messages.success(request, f"ƒê√£ c·∫≠p nh·∫≠t m·∫´u ch·∫•m cho {bai_thi.tenBaiThi}: {len(section_map)} m·ª•c l·ªõn, {created_items} m·ª•c nh·ªè.")
            return redirect(request.path)

        # (C√°c action kh√°c nh∆∞ create_ct/create_vt/create_bt/config_time_rules c√≥ th·ªÉ b·ªï sung sau)
        messages.error(request, "H√†nh ƒë·ªông ch∆∞a ƒë∆∞·ª£c h·ªó tr·ª£.")
        return redirect(request.path)

    # GET: hi·ªÉn th·ªã trang t·ªï ch·ª©c
    cuoc_this = (
        CuocThi.objects
        .prefetch_related(
            Prefetch("vong_thi", queryset=VongThi.objects.prefetch_related(
                Prefetch("bai_thi", queryset=BaiThi.objects.prefetch_related("time_rules", "template_sections__items"))
            ))
        )
        .order_by("ma")
    )
    return render(request, "organize/index.html", {"cuoc_this": cuoc_this})