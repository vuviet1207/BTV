from django.shortcuts import render, redirect
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.db.models import Prefetch
from django.http import JsonResponse, QueryDict
import json
from .models import CuocThi, VongThi, BaiThi, BaiThiTimeRule, BaiThiTemplateSection, BaiThiTemplateItem, GiamKhao, GiamKhaoBaiThi

from .models import (
    CuocThi, VongThi, BaiThi,
    BaiThiTimeRule,
    BaiThiTemplateSection, BaiThiTemplateItem,
    GiamKhao, GiamKhaoBaiThi,
)


# ============================================================
# /organize/  (màn hình quản lý CT → VT → BT)
# ============================================================
@require_http_methods(["GET", "POST"])
def organize_view(request, ct_id=None):
    if request.method == "POST":
        action = request.POST.get("action")
        try:
            # --------------------------------------------------
            # Bật/Tắt cuộc thi
            # --------------------------------------------------
            if action == "toggle_ct":
                ct_pk = request.POST.get("cuocThi_id")
                if not ct_pk:
                    messages.error(request, "Thiếu mã cuộc thi.")
                    return redirect(request.path)
                ct = CuocThi.objects.get(id=ct_pk)
                new_state = (request.POST.get("trangThai") == "on")
                ct.trangThai = new_state
                ct.save(update_fields=["trangThai"])
                messages.success(request, f"{ct.ma}: đã {'bật' if new_state else 'tắt'}.")
                return redirect(request.path)

            # --------------------------------------------------
            # Tạo cuộc thi
            # --------------------------------------------------
            if action == "create_ct":
                ten = (request.POST.get("tenCuocThi") or "").strip()
                trang_thai = request.POST.get("trangThai") == "on"
                if not ten:
                    messages.error(request, "Vui lòng nhập tên cuộc thi.")
                else:
                    ct = CuocThi(tenCuocThi=ten, trangThai=trang_thai)
                    ct.save()  # auto-gen ct.ma
                    messages.success(request, f"Tạo cuộc thi {ct.ma} thành công.")
                return redirect(request.path)

            # --------------------------------------------------
            # Tạo vòng thi
            # --------------------------------------------------
            if action == "create_vt":
                ct_id = request.POST.get("cuocThi_id")
                ten_vt = (request.POST.get("tenVongThi") or "").strip()
                if not ct_id or not ten_vt:
                    messages.error(request, "Vui lòng chọn cuộc thi và nhập tên vòng thi.")
                else:
                    cuoc_thi = CuocThi.objects.get(id=ct_id)
                    vt = VongThi(tenVongThi=ten_vt, cuocThi=cuoc_thi)
                    vt.save()  # auto-gen vt.ma
                    messages.success(request, f"Tạo vòng thi {vt.ma} trong {cuoc_thi.ma} thành công.")
                return redirect(request.path)

            # --------------------------------------------------
            # Đổi tên vòng thi
            # --------------------------------------------------
            if action == "rename_vt":
                vt_id = request.POST.get("vongThi_id")
                new_name = (request.POST.get("tenVongThi") or "").strip()

                if not vt_id:
                    messages.error(request, "Thiếu ID vòng thi.")
                    return redirect(request.path)

                try:
                    vt = VongThi.objects.get(id=vt_id)
                except VongThi.DoesNotExist:
                    messages.error(request, "Vòng thi không tồn tại.")
                    return redirect(request.path)

                if not new_name:
                    messages.error(request, "Tên vòng thi không được để trống.")
                    return redirect(request.path)

                old_name = vt.tenVongThi
                vt.tenVongThi = new_name
                vt.save(update_fields=["tenVongThi"])

                messages.success(
                    request,
                    f"Đã đổi tên vòng thi “{old_name}” → “{new_name}”."
                )
                return redirect(request.path)
            # --------------------------------------------------
            # Tạo bài thi
            # --------------------------------------------------
            if action == "create_bt":
                vt_id = request.POST.get("vongThi_id")
                ten_bt = (request.POST.get("tenBaiThi") or "").strip()
                method = request.POST.get("phuongThucCham") or "POINTS"
                max_diem = request.POST.get("cachChamDiem")
                judge_code = (request.POST.get("judge_id") or "").strip()  # CHỈ 1 mã, có thể rỗng

                if not vt_id or not ten_bt:
                    messages.error(request, "Vui lòng chọn vòng thi và nhập tên bài thi.")
                    return redirect(request.path)

                # ❌ BỎ ràng buộc bắt buộc phải chọn giám khảo
                # if not judge_ids:
                #     messages.error(request, "Vui lòng chọn giám khảo chấm cho bài thi.")
                #     return redirect(request.path)

                if method == "POINTS" and not max_diem:
                    messages.error(request, "Vui lòng nhập điểm tối đa cho phương thức thang điểm.")
                    return redirect(request.path)

                vong_thi = VongThi.objects.get(id=vt_id)
                bt = BaiThi.objects.create(
                    tenBaiThi=ten_bt,
                    vongThi=vong_thi,
                    phuongThucCham=method,
                    cachChamDiem=int(max_diem) if (method == "POINTS" and max_diem) else 0,
                )

                # ✅ Nếu có chọn giám khảo thì mới tạo phân công
                if judge_code:
                    gk = GiamKhao.objects.filter(maNV=judge_code, role="JUDGE").first()
                    if not gk:
                        bt.delete()
                        messages.error(request, f"Giám khảo {judge_code} không hợp lệ.")
                        return redirect(request.path)
                    GiamKhaoBaiThi.objects.get_or_create(giamKhao=gk, baiThi=bt)

                messages.success(
                    request,
                    f"Tạo bài thi “{bt.tenBaiThi}” trong vòng “{vong_thi.tenVongThi}” thành công."
                )
                return redirect(request.path)


            # --------------------------------------------------
            # Lưu cấu hình thang thời gian (từ popup TIME)
            # --------------------------------------------------
            if action == "config_time_rules":
                import json
                btid = request.POST.get("baiThi_id")
                raw = request.POST.get("time_rules_json") or "[]"

                try:
                    bt = BaiThi.objects.get(id=btid)
                except BaiThi.DoesNotExist:
                    messages.error(request, "Bài thi không tồn tại.")
                    return redirect(request.path)

                if bt.phuongThucCham != "TIME":
                    messages.error(request, "Bài thi này không dùng phương thức chấm theo thời gian.")
                    return redirect(request.path)

                try:
                    rows = json.loads(raw)
                except Exception:
                    messages.error(request, "Dữ liệu cấu hình không hợp lệ.")
                    return redirect(request.path)

                # Validate & lưu
                cleaned = []
                for r in rows:
                    try:
                        s = int(r.get("start", 0))
                        e = int(r.get("end", 0))
                        sc = int(r.get("score", 0))
                    except Exception:
                        continue
                    if s < 0 or e < 0 or e < s:
                        continue
                    cleaned.append((s, e, sc))

                from django.db import transaction
                with transaction.atomic():
                    bt.time_rules.all().delete()
                    BaiThiTimeRule.objects.bulk_create([
                        BaiThiTimeRule(baiThi=bt, start_seconds=s, end_seconds=e, score=sc)
                        for (s, e, sc) in cleaned
                    ])

                messages.success(request, f"Đã lưu {len(cleaned)} dòng thang thời gian cho {bt.ma}.")
                return redirect(request.path)

            # --------------------------------------------------
            # Import TEMPLATE (Excel) → tạo Section/Item
            # --------------------------------------------------
            if action == "config_template_upload":
                btid = request.POST.get("baiThi_id")
                f = request.FILES.get("template_file")

                if not btid or not f:
                    messages.error(request, "Thiếu bài thi hoặc tệp Excel.")
                    return redirect(request.path)

                try:
                    bt = BaiThi.objects.get(id=btid)
                except BaiThi.DoesNotExist:
                    messages.error(request, "Bài thi không tồn tại.")
                    return redirect(request.path)

                if bt.phuongThucCham != "TEMPLATE":
                    messages.error(request, "Bài thi này không dùng phương thức chấm theo mẫu.")
                    return redirect(request.path)

                # Đọc Excel
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(f, data_only=True)
                    ws = wb[wb.sheetnames[0]]
                except Exception:
                    messages.error(request, "Không đọc được file Excel. Vui lòng dùng .xlsx hợp lệ.")
                    return redirect(request.path)

                # Tìm header linh hoạt: "Danh Mục 1 | Danh Mục 2 | Điểm"
                header_row_idx = None
                col_section = col_item = col_max = None

                def norm(v):
                    return str(v).strip().lower() if v is not None else ""

                for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
                    texts = [norm(c) for c in row]
                    for cidx, t in enumerate(texts):
                        if t in ("danh mục 1", "mục lớn", "section"):
                            col_section = cidx
                        elif t in ("danh mục 2", "mục nhỏ", "item", "nội dung"):
                            col_item = cidx
                        elif ("điểm" in t) and ("chấm" not in t):
                            col_max = cidx
                    if col_section is not None and col_item is not None and col_max is not None:
                        header_row_idx = ridx
                        break

                if header_row_idx is None:
                    messages.error(request, "Không tìm thấy dòng tiêu đề (Danh Mục 1 / Danh Mục 2 / Điểm).")
                    return redirect(request.path)

                # Gom dữ liệu từ hàng ngay sau header
                rows = []
                for r in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    v_section = r[col_section] if len(r) > col_section else None
                    v_item    = r[col_item]    if len(r) > col_item    else None
                    v_max_raw = r[col_max]     if len(r) > col_max     else None

                    s = (str(v_section or "").strip())
                    i = (str(v_item or "").strip())
                    if not s and not i and v_max_raw in (None, ""):
                        continue

                    # ép max score
                    mx = 0
                    if isinstance(v_max_raw, (int, float)):
                        mx = int(v_max_raw)
                    else:
                        try:
                            mx = int(str(v_max_raw).strip()) if v_max_raw not in (None, "") else 0
                        except Exception:
                            mx = 0

                    if not s and not i:
                        continue
                    if s and i:
                        rows.append((s, i, mx, ""))

                if not rows:
                    messages.error(request, "Không tìm thấy dòng dữ liệu nào dưới tiêu đề (Danh Mục 1/2, Điểm).")
                    return redirect(request.path)

                # ====== LƯU VÀO DB (không map theo title để tránh KeyError) ======
                from django.db import transaction

                def _clean_title(s: str) -> str:
                    # gom mọi xuống dòng/khoảng trắng thành 1 space cho đẹp
                    return " ".join((s or "").split())

                with transaction.atomic():
                    bt.template_sections.all().delete()

                    # 1) Lấy danh sách section theo thứ tự xuất hiện (dùng để map)
                    section_order: list[str] = []
                    section_index: dict[str, int] = {}
                    for (sect, _item, _mx, _note) in rows:
                        key = sect  # giữ nguyên khóa gốc từ Excel
                        if key not in section_index:
                            section_index[key] = len(section_order) + 1  # STT bắt đầu từ 1
                            section_order.append(key)

                    # 2) Tạo Sections (hiển thị: đã làm sạch + cắt 255 ký tự)
                    section_objs = []
                    for name in section_order:
                        display = _clean_title(name)[:255]  # nếu model max_length khác 255 thì đổi số này
                        section_objs.append(
                            BaiThiTemplateSection(baiThi=bt, stt=section_index[name], title=display)
                        )
                    BaiThiTemplateSection.objects.bulk_create(section_objs)

                    # 3) Lấy lại theo STT & map bằng THỨ TỰ (không map theo title nữa)
                    created_list = list(
                        BaiThiTemplateSection.objects.filter(baiThi=bt).order_by("stt")
                    )
                    section_by_key = {
                        section_order[i]: created_list[i] for i in range(len(section_order))
                    }

                    # 4) Tạo Items
                    counters = {k: 0 for k in section_order}
                    items_to_create = []
                    for (sect, item, mx, note) in rows:
                        # nếu vì lý do nào đó sect không có trong map → bỏ qua để không crash
                        if sect not in section_by_key:
                            continue
                        counters[sect] += 1
                        # ép max_score an toàn
                        if isinstance(mx, (int, float)):
                            mx_val = int(mx)
                        else:
                            try:
                                mx_val = int(str(mx).strip())
                            except Exception:
                                mx_val = 0
                        items_to_create.append(
                            BaiThiTemplateItem(
                                section=section_by_key[sect],
                                stt=counters[sect],
                                content=(item or "").strip(),
                                max_score=mx_val,
                                note=(note or None),
                            )
                        )
                    BaiThiTemplateItem.objects.bulk_create(items_to_create)


                messages.success(
                    request,
                    f"Đã import {len(section_order)} mục lớn và {len(items_to_create)} mục nhỏ cho {bt.ma}."
                )
                return redirect(request.path)

            # --------------------------------------------------
            # PREVIEW import thang thời gian (Excel A..E, dữ liệu từ hàng 3)
            # Trả JSON để JS đổ lên popup (KHÔNG lưu DB)
            # --------------------------------------------------
            if action == "time_upload_preview":
                btid = request.POST.get("baiThi_id")
                f = request.FILES.get("time_file")
                if not btid or not f:
                    return JsonResponse({"ok": False, "error": "Thiếu bài thi hoặc tệp Excel."}, status=400)

                try:
                    bt = BaiThi.objects.get(id=btid)
                except BaiThi.DoesNotExist:
                    return JsonResponse({"ok": False, "error": "Bài thi không tồn tại."}, status=404)

                if bt.phuongThucCham != "TIME":
                    return JsonResponse({"ok": False, "error": "Bài thi này không dùng phương thức chấm theo thời gian."}, status=400)

                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(f, data_only=True)
                    ws = wb[wb.sheetnames[0]]
                except Exception:
                    return JsonResponse({"ok": False, "error": "Không đọc được file .xlsx."}, status=400)

                def as_int(x):
                    if x is None or str(x).strip() == "":
                        raise ValueError("empty")
                    return int(float(x))  # xử lý 10.0

                rows = []
                r = 3  # dữ liệu bắt đầu tại hàng 3
                while True:
                    sm = ws.cell(row=r, column=1).value  # A: phút bắt đầu
                    ss = ws.cell(row=r, column=2).value  # B: giây  bắt đầu
                    em = ws.cell(row=r, column=3).value  # C: phút kết thúc
                    es = ws.cell(row=r, column=4).value  # D: giây  kết thúc
                    sc = ws.cell(row=r, column=5).value  # E: điểm

                    # nếu cả 5 ô đều trống → coi như hết dữ liệu
                    if all(v in (None, "", " ") for v in (sm, ss, em, es, sc)):
                        break

                    try:
                        smi = as_int(sm); ssi = as_int(ss)
                        emi = as_int(em); esi = as_int(es)
                        sci = as_int(sc)
                    except Exception:
                        r += 1
                        continue

                    # validate cơ bản
                    if ssi < 0 or ssi > 59 or esi < 0 or esi > 59:
                        r += 1; continue
                    if smi < 0 or emi < 0 or sci < 0:
                        r += 1; continue

                    start = smi * 60 + ssi
                    end   = emi * 60 + esi
                    if end <= start:
                        r += 1; continue

                    rows.append({"start": start, "end": end, "score": sci})
                    r += 1

                if not rows:
                    return JsonResponse({"ok": False, "error": "Không đọc được dòng hợp lệ nào (từ hàng 3)."}, status=400)

                return JsonResponse({"ok": True, "rows": rows})
            
            if action == "delete_vt":
                vtid = request.POST.get("vongThi_id")
                try:
                    vt = VongThi.objects.get(id=vtid)
                except VongThi.DoesNotExist:
                    return JsonResponse({"ok": False, "error": "Vòng thi không tồn tại."}, status=404)

                name = vt.tenVongThi
                vt.delete()  # sẽ cascade xoá các bài thi + cấu hình liên quan

                return JsonResponse({"ok": True, "message": f"Đã xoá vòng thi {name}."})

            if action == "delete_bt":
                btid = request.POST.get("baiThi_id")
                try:
                    bt = BaiThi.objects.get(id=btid)
                except BaiThi.DoesNotExist:
                    return JsonResponse({"ok": False, "error": "Bài thi không tồn tại."}, status=404)

                name = bt.tenBaiThi
                bt.delete()

                return JsonResponse({"ok": True, "message": f"Đã xoá bài thi {name}."})

            # --------------------------------------------------
            # Action không khớp
            # --------------------------------------------------
            # Update judge assignments for a baiThi (AJAX)
            # Accept either a regular form POST with action=update_assignments or a JSON POST
            # (some clients send Content-Type: application/json; charset=utf-8 so check substring)
            if action == "update_assignments" or (request.content_type and 'application/json' in request.content_type):
                # Accept form-encoded or JSON body
                payload = {}
                try:
                    ct = (request.content_type or '').lower()
                    if 'application/json' in ct:
                        raw = request.body or b''
                        if not raw:
                            # empty body — treat as empty payload
                            payload = {}
                        else:
                            try:
                                payload = json.loads(raw.decode('utf-8'))
                            except Exception as e:
                                body_snippet = raw.decode('utf-8', errors='replace')[:500]
                                # log minimal info for debugging (server console)
                                print(f"[organize] JSON parse error: {e}; content_type={request.content_type}; body_snippet={body_snippet!r}")
                                return JsonResponse({"ok": False, "message": f"Invalid payload (JSON parse error). content_type={request.content_type}; body_snippet={body_snippet[:200]}"}, status=400)
                    else:
                        # form-encoded (QueryDict)
                        payload = request.POST
                except Exception as e:
                    print(f"[organize] payload parsing unexpected error: {e}")
                    return JsonResponse({"ok": False, "message": "Invalid payload"}, status=400)

                # baiThi id
                btid = None
                if isinstance(payload, dict):
                    btid = payload.get('baiThi_id')
                else:
                    # QueryDict (request.POST)
                    btid = payload.get('baiThi_id') or request.POST.get('baiThi_id')

                # judges: support JSON array, comma-separated string, or repeated form fields
                judges = []
                if isinstance(payload, QueryDict):
                    # request.POST: getlist returns all repeated 'judges' values
                    judges = payload.getlist('judges') or []
                else:
                    # payload is dict from JSON parsing
                    raw_judges = payload.get('judges') if isinstance(payload, dict) else None
                    if isinstance(raw_judges, list):
                        judges = raw_judges
                    elif isinstance(raw_judges, str):
                        # comma-separated
                        judges = [s.strip() for s in raw_judges.split(',') if s.strip()]
                    else:
                        judges = []

                if not btid:
                    return JsonResponse({"ok": False, "message": "Missing baiThi_id"}, status=400)

                try:
                    bt = BaiThi.objects.get(id=btid)
                except BaiThi.DoesNotExist:
                    return JsonResponse({"ok": False, "message": "BaiThi not found"}, status=404)

                # Current assigned set
                current = set(g.giamKhao.maNV for g in bt.giam_khao_duoc_chi_dinh.all())
                newset = set(judges)

                to_add = newset - current
                to_remove = current - newset

                from django.db import transaction
                with transaction.atomic():
                    # remove
                    if to_remove:
                        GiamKhaoBaiThi.objects.filter(
                            baiThi=bt,
                            giamKhao__maNV__in=list(to_remove)
                        ).delete()
                    # add (only if giamkhao exists)
                    for ma in to_add:
                        try:
                            gk = GiamKhao.objects.get(maNV=ma)
                            GiamKhaoBaiThi.objects.create(giamKhao=gk, baiThi=bt)
                        except GiamKhao.DoesNotExist:
                            # skip unknown judge codes
                            continue

                return JsonResponse({"ok": True, "message": "Assignments updated", "added": list(to_add), "removed": list(to_remove)})

            messages.error(request, "Hành động không hợp lệ.")
            return redirect(request.path)

        except CuocThi.DoesNotExist:
            messages.error(request, "Cuộc thi không tồn tại.")
            return redirect(request.path)
        except VongThi.DoesNotExist:
            messages.error(request, "Vòng thi không tồn tại.")
            return redirect(request.path)
        except ValueError:
            messages.error(request, "Giá trị điểm tối đa không hợp lệ.")
            return redirect(request.path)

    # ==========================
    # GET: render màn hình Organize
    # ==========================
    base_qs = CuocThi.objects.prefetch_related(
        Prefetch(
            "vong_thi",
            queryset=VongThi.objects.prefetch_related(
                "bai_thi__time_rules",
                "bai_thi__template_sections__items",
                Prefetch(
                    "bai_thi__giam_khao_duoc_chi_dinh",
                    queryset=GiamKhaoBaiThi.objects.select_related("giamKhao"),
                    to_attr="assignments"
                ),
            )
        )
    ).order_by("-id")

    if ct_id:
        base_qs = base_qs.filter(id=ct_id)
    judges = list(GiamKhao.objects.filter(role="JUDGE").all())

    # Prepare JSON-serializable payload for the template's `json_script` helper.
    # Template expects `judges_payload` and uses it as `ALL_JUDGES` in JS.
    # The frontend expects items with `code` and `name` properties
    judges_payload = [
        {
            "code": getattr(g, "maNV", ""),
            "name": getattr(g, "hoTen", ""),
            "email": getattr(g, "email", ""),
        }
        for g in judges
    ]

    return render(request, "organize/index.html", {"cuoc_this": base_qs, "judges": judges, "judges_payload": judges_payload})


# ============================================================
# /competitions/  (danh sách cuộc thi đơn giản)
# ============================================================
@require_http_methods(["GET", "POST"])
def competition_list_view(request):
    if request.method == "POST":
        action = request.POST.get("action")
        try:
            if action == "create":
                ten = (request.POST.get("tenCuocThi") or "").strip()
                trang_thai = request.POST.get("trangThai") == "on"
                if not ten:
                    messages.error(request, "Vui lòng nhập tên cuộc thi.")
                else:
                    ct = CuocThi(tenCuocThi=ten, trangThai=trang_thai)
                    ct.save()
                    messages.success(request, f"Đã tạo {ct.ma}.")
                return redirect(request.path)

            if action == "update":
                ct_id = request.POST.get("id")
                ten = (request.POST.get("tenCuocThi") or "").strip()
                trang_thai = request.POST.get("trangThai") == "on"
                ct = CuocThi.objects.get(id=ct_id)
                if not ten:
                    messages.error(request, "Tên cuộc thi không được rỗng.")
                else:
                    ct.tenCuocThi = ten
                    ct.trangThai = trang_thai
                    ct.save(update_fields=["tenCuocThi", "trangThai"])
                    messages.success(request, f"Đã cập nhật {ct.ma}.")
                return redirect(request.path)

            if action == "delete":
                ct_id = request.POST.get("id")
                ct = CuocThi.objects.get(id=ct_id)
                code = ct.ma
                ct.delete()
                messages.success(request, f"Đã xoá {code}.")
                return redirect(request.path)

            messages.error(request, "Hành động không hợp lệ.")
            return redirect(request.path)

        except CuocThi.DoesNotExist:
            messages.error(request, "Cuộc thi không tồn tại.")
            return redirect(request.path)

    # GET
    items = CuocThi.objects.order_by("-id")
    return render(request, "organize/competitions.html", {"items": items})
